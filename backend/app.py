from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import jieba
import jieba.analyse
from collections import Counter
import re
import json
import os
import pandas as pd
import io
from werkzeug.utils import secure_filename

# 导入情感标签定义
try:
    from sentiment_labels import (
        SENTIMENT_LABEL_LIST, LABEL_TO_CN, CN_TO_LABEL, 
        LABEL_TO_COLOR, convert_old_label_to_new, is_valid_sentiment_label
    )
    SENTIMENT_LABELS_AVAILABLE = True
except ImportError:
    SENTIMENT_LABELS_AVAILABLE = False
    print("警告: 情感标签模块未找到，将使用默认三分类")
    # 默认映射（兼容旧代码）
    SENTIMENT_LABEL_LIST = ['negative', 'neutral', 'positive']
    LABEL_TO_CN = {'negative': '负面', 'neutral': '中性', 'positive': '正面'}
    CN_TO_LABEL = {'负面': 'negative', '中性': 'neutral', '正面': 'positive'}
    LABEL_TO_COLOR = {'negative': '#f44336', 'neutral': '#ff9800', 'positive': '#4caf50'}
    def convert_old_label_to_new(label, text=''): return label
    def is_valid_sentiment_label(label): return label in SENTIMENT_LABEL_LIST

# 导入ML分析模块
try:
    from ml_analyzer import create_ml_analyzer, MLAnalyzer
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False
    print("警告: ML分析模块未找到")

# 导入AI分析模块
try:
    from ai_analyzer import ai_analyzer
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False
    print("警告: AI分析模块未找到，将仅使用ML模型")

# 导入隐私保护模块
try:
    from privacy_utils import privacy_utils
    PRIVACY_AVAILABLE = True
except ImportError:
    PRIVACY_AVAILABLE = False
    print("警告: 隐私保护模块未找到，将不进行敏感信息检测")

app = Flask(__name__)
CORS(app)

# 初始化jieba
jieba.initialize()

# 人机协同配置
# 置信度区间配置
HUMAN_REVIEW_MIN_THRESHOLD = float(os.getenv('HUMAN_REVIEW_MIN_THRESHOLD', '0.5'))  # 需要人工复核的最小置信度
HUMAN_REVIEW_MAX_THRESHOLD = float(os.getenv('HUMAN_REVIEW_MAX_THRESHOLD', '0.85'))  # 需要人工复核的最大置信度（超过此值自动采纳）
# 是否自动将人工复核结果添加到训练数据
AUTO_ADD_TO_TRAINING = os.getenv('AUTO_ADD_TO_TRAINING', 'true').lower() == 'true'
# 是否自动重训练模型（当添加新数据时）
AUTO_RETRAIN_ON_UPDATE = os.getenv('AUTO_RETRAIN_ON_UPDATE', 'false').lower() == 'true'

# 存储人工复核结果（实际应用中应使用数据库）
human_review_results = {}

# 加载训练数据
def load_training_data():
    """
    加载训练数据
    从 training_data.json 文件加载基础训练数据
    自动将旧的三分类标签转换为新的五分类标签
    
    Returns:
        训练数据列表，格式为 [(text, label), ...]
    """
    training_data = []
    
    # 从 training_data.json 加载基础训练数据
    training_data_file = "training_data.json"
    try:
        if os.path.exists(training_data_file):
            with open(training_data_file, 'r', encoding='utf-8') as f:
                base_data = json.load(f)
                # 转换旧标签为新标签
                converted_count = 0
                for item in base_data:
                    old_label = item.get('label', 'neutral')
                    text = item.get('text', '')
                    # 如果是旧标签，转换为新标签
                    if old_label in ['positive', 'negative', 'neutral'] and SENTIMENT_LABELS_AVAILABLE:
                        new_label = convert_old_label_to_new(old_label, text)
                        if new_label != old_label:
                            converted_count += 1
                        training_data.append((text, new_label))
                    else:
                        # 已经是新标签或无效标签
                        if SENTIMENT_LABELS_AVAILABLE and is_valid_sentiment_label(old_label):
                            training_data.append((text, old_label))
                        else:
                            # 无效标签，使用默认值
                            training_data.append((text, 'neutral'))
                print(f"[数据加载] 从 {training_data_file} 加载了 {len(base_data)} 条基础数据，转换了 {converted_count} 条标签")
        else:
            print(f"[数据加载] 警告: {training_data_file} 文件不存在，将使用空数据")
    except Exception as e:
        print(f"[数据加载] 加载 {training_data_file} 失败: {e}")
    
    # 去重（基于文本内容）
    seen = set()
    unique_data = []
    for text, label in training_data:
        if text not in seen:
            seen.add(text)
            unique_data.append((text, label))
    
    print(f"[数据加载] 训练数据加载完成，共 {len(unique_data)} 条（去重后）")
    return unique_data

def save_training_data(training_data_list):
    """
    保存训练数据到 training_data.json 文件
    
    Args:
        training_data_list: 训练数据列表，格式为 [{"text": "...", "label": "..."}, ...]
    """
    training_data_file = "training_data.json"
    try:
        with open(training_data_file, 'w', encoding='utf-8') as f:
            json.dump(training_data_list, f, ensure_ascii=False, indent=2)
        print(f"[数据保存] 成功保存 {len(training_data_list)} 条训练数据到 {training_data_file}")
        return True
    except Exception as e:
        print(f"[数据保存] 保存 {training_data_file} 失败: {e}")
        return False

def add_to_training_data(text, label):
    """
    将新的训练数据添加到 training_data.json
    
    Args:
        text: 评论文本
        label: 情感标签 (支持五分类：strongly_negative/weakly_negative/neutral/weakly_positive/strongly_positive)
              也兼容旧标签 (positive/negative/neutral)，会自动转换
    
    Returns:
        bool: 是否成功添加（如果已存在则返回False）
    """
    training_data_file = "training_data.json"
    
    # 如果是旧标签，转换为新标签
    if SENTIMENT_LABELS_AVAILABLE and label in ['positive', 'negative', 'neutral']:
        label = convert_old_label_to_new(label, text)
        print(f"[数据添加] 标签已转换: {label}")
    
    # 验证标签有效性
    if SENTIMENT_LABELS_AVAILABLE and not is_valid_sentiment_label(label):
        print(f"[数据添加] 警告: 无效的标签 '{label}'，使用默认标签 'neutral'")
        label = 'neutral'
    
    # 加载现有数据
    existing_data = []
    if os.path.exists(training_data_file):
        try:
            with open(training_data_file, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
        except Exception as e:
            print(f"[数据添加] 加载现有数据失败: {e}")
            existing_data = []
    
    # 检查是否已存在（基于文本内容去重）
    for item in existing_data:
        if item.get('text', '').strip() == text.strip():
            print(f"[数据添加] 数据已存在，跳过: {text[:50]}...")
            return False
    
    # 添加新数据
    new_item = {
        "text": text.strip(),
        "label": label
    }
    existing_data.append(new_item)
    
    # 保存到文件
    if save_training_data(existing_data):
        label_cn = LABEL_TO_CN.get(label, label) if SENTIMENT_LABELS_AVAILABLE else label
        print(f"[数据添加] 成功添加新训练数据: {text[:50]}... (标签: {label_cn})")
        return True
    else:
        return False

# 加载训练数据
training_data = load_training_data()

# 初始化ML分析器
if ML_AVAILABLE:
    ml_analyzer = create_ml_analyzer(training_data)
else:
    ml_analyzer = None
    print("警告: ML分析器初始化失败")

def analyze_sentiment_ml(text):
    """
    使用机器学习模型分析文本情感
    
    Args:
        text: 待分析的文本
        
    Returns:
        情感分析结果
    """
    if ml_analyzer is None:
        raise ValueError("ML分析器未初始化")
    return ml_analyzer.analyze_sentiment(text)

def check_human_review_needed(confidence):
    """
    根据置信度判断是否需要人工复核
    
    规则：
    - confidence < 0.5: 标记为"无效"或"待观察"
    - 0.5 <= confidence <= 0.85: 需要人工复核
    - confidence > 0.85: 系统自动采纳
    
    Args:
        confidence: 模型输出的置信度
        
    Returns:
        dict: {
            "needs_review": bool,  # 是否需要人工复核
            "status": str,  # "invalid" | "needs_review" | "auto_accepted"
            "reason": str   # 状态说明
        }
    """
    if confidence < HUMAN_REVIEW_MIN_THRESHOLD:
        return {
            "needs_review": False,
            "status": "invalid",
            "reason": f"置信度过低（{confidence:.2f} < {HUMAN_REVIEW_MIN_THRESHOLD}），标记为无效/待观察"
        }
    elif confidence > HUMAN_REVIEW_MAX_THRESHOLD:
        return {
            "needs_review": False,
            "status": "auto_accepted",
            "reason": f"置信度很高（{confidence:.2f} > {HUMAN_REVIEW_MAX_THRESHOLD}），系统自动采纳"
        }
    else:
        return {
            "needs_review": True,
            "status": "needs_review",
            "reason": f"置信度在复核区间内（{HUMAN_REVIEW_MIN_THRESHOLD} <= {confidence:.2f} <= {HUMAN_REVIEW_MAX_THRESHOLD}），需要人工复核"
        }

def analyze_sentiment(text, use_ai=None):
    """
    智能情感分析：优先使用AI，失败时回退到ML
    模型负责对全部评论进行初筛，并输出其置信度
    
    Args:
        text: 评论文本
        use_ai: 是否使用AI，None表示自动选择
        
    Returns:
        情感分析结果，包含confidence和human_review_needed字段
    """
    # 确定分析模式
    if use_ai is None:
        # 从请求参数或环境变量获取
        try:
            from flask import has_request_context
            if has_request_context() and request:
                analysis_mode = request.json.get('analysis_mode', 'auto') if request.json else 'auto'
            else:
                analysis_mode = 'auto'
        except:
            analysis_mode = 'auto'
        
        if analysis_mode == 'auto':
            # auto模式：如果AI可用且启用，则使用AI
            use_ai = AI_AVAILABLE and ai_analyzer.ai_enabled if AI_AVAILABLE else False
        else:
            use_ai = (analysis_mode == 'ai' or analysis_mode == 'hybrid')
    else:
        # 如果use_ai已明确指定，尝试获取analysis_mode用于混合模式判断
        try:
            from flask import has_request_context
            if has_request_context() and request:
                analysis_mode = request.json.get('analysis_mode', 'hybrid' if use_ai else 'ml') if request.json else ('hybrid' if use_ai else 'ml')
            else:
                analysis_mode = 'hybrid' if use_ai else 'ml'
        except:
            analysis_mode = 'hybrid' if use_ai else 'ml'
    
    result = None
    
    # 尝试使用AI分析
    if use_ai and AI_AVAILABLE:
        result = ai_analyzer.analyze_sentiment_with_ai(text)
        print(result, '-----------------airesult-----------------')
    
    # 如果AI分析失败或未启用，使用ML模型
    if result is None:
        result = analyze_sentiment_ml(text)
        print(result, '-----mlresult------')
    else:
        # 混合模式：如果AI置信度较低，结合ML结果
        if analysis_mode == 'hybrid' and result.get('confidence', 1.0) < 0.8:
            ml_result = analyze_sentiment_ml(text)
            # 如果ML结果置信度更高，使用ML结果
            if ml_result.get('confidence', 0) > result.get('confidence', 0):
                result = ml_result
                result['method'] = 'hybrid_ml'
            else:
                result['method'] = 'hybrid_ai'
    
    # 人机协同校验：根据置信度判断状态
    confidence = result.get('confidence', 0.5)
    review_status = check_human_review_needed(confidence)
    result['human_review_needed'] = review_status['needs_review']
    result['review_status'] = review_status['status']  # "invalid" | "needs_review" | "auto_accepted"
    result['review_reason'] = review_status['reason']
    result['confidence_thresholds'] = {
        "min": HUMAN_REVIEW_MIN_THRESHOLD,
        "max": HUMAN_REVIEW_MAX_THRESHOLD
    }
    
    return result

def extract_keywords(text, topK=10, use_ai=None):
    """
    智能关键词提取：优先使用AI，失败时使用传统方法
    
    Args:
        text: 评论文本
        topK: 返回关键词数量
        use_ai: 是否使用AI，None表示自动选择
        
    Returns:
        关键词列表
    """
    # 尝试使用AI提取
    if use_ai is None:
        # 尝试从请求上下文获取，如果没有则使用默认值
        try:
            from flask import has_request_context
            if has_request_context():
                analysis_mode = request.json.get('analysis_mode', 'auto') if request else 'auto'
            else:
                analysis_mode = 'auto'
        except:
            analysis_mode = 'auto'
        use_ai = (analysis_mode == 'ai' or analysis_mode == 'hybrid') if AI_AVAILABLE else False
    else:
        use_ai = use_ai and AI_AVAILABLE
    
    if use_ai and AI_AVAILABLE:
        ai_keywords = ai_analyzer.extract_keywords_with_ai(text, topK)
        if ai_keywords:
            return ai_keywords
    
    # 使用传统TF-IDF方法
    if ml_analyzer is not None:
        return ml_analyzer.extract_keywords(text, topK)
    else:
        keywords = jieba.analyse.extract_tags(text, topK=topK, withWeight=True)
        return [{"word": word, "weight": float(weight)} for word, weight in keywords]

def calculate_satisfaction_score(comments):
    """计算满意度评分"""
    if not comments:
        return {"score": 0, "level": "无数据"}
    
    total_score = 0
    
    # 五分类评分体系
    if SENTIMENT_LABELS_AVAILABLE:
        sentiment_scores = {
            "strongly_positive": 5,
            "weakly_positive": 4,
            "neutral": 3,
            "weakly_negative": 2,
            "strongly_negative": 1
        }
    else:
        # 兼容旧的三分类
        sentiment_scores = {
            "positive": 5,
            "neutral": 3,
            "negative": 1
        }
    
    for comment in comments:
        sentiment = analyze_sentiment(comment["text"])
        sentiment_type = sentiment["sentiment"]
        total_score += sentiment_scores.get(sentiment_type, 3)
    
    avg_score = total_score / len(comments)
    
    # 评分等级
    if avg_score >= 4.5:
        level = "非常满意"
    elif avg_score >= 3.5:
        level = "满意"
    elif avg_score >= 2.5:
        level = "一般"
    else:
        level = "不满意"
    
    return {
        "score": round(avg_score, 2),
        "level": level,
        "total_comments": len(comments)
    }

@app.route('/api/analyze', methods=['POST'])
def analyze_comment():
    """分析单个评论"""
    data = request.json
    original_text = data.get('text', '')
    analysis_mode = data.get('analysis_mode', 'auto')  # auto, ai, ml, hybrid
    industry = data.get('industry', None)  # 行业类型（如'adult_products'）
    
    if not original_text:
        return jsonify({"error": "评论内容不能为空"}), 400
    
    # 隐私保护：检测敏感信息
    sensitive_info = {}
    privacy_info = {
        "sensitive_detected": False,
        "local_analysis_used": False,
        "desensitization_applied": False
    }
    
    if PRIVACY_AVAILABLE:
        sensitive_info = privacy_utils.detect_sensitive_info(original_text)
        privacy_info["sensitive_detected"] = any(sensitive_info.values())
        
        # 判断是否应该使用本地分析（不上传AI）
        should_use_local = privacy_utils.should_use_local_analysis(original_text, industry)
        
        # 如果包含敏感信息或特殊行业，强制使用本地分析
        if should_use_local and analysis_mode != 'ai':
            analysis_mode = 'ml'  # 强制使用ML，不上传AI
            privacy_info["local_analysis_used"] = True
    
    # 根据模式选择分析方法
    # auto模式：传递None让analyze_sentiment自动决定
    # ai/hybrid模式：强制使用AI（但如果检测到敏感信息，会降级到ML）
    # ml模式：强制使用ML
    if analysis_mode == 'auto':
        use_ai = None  # 让analyze_sentiment自动决定
    elif analysis_mode == 'ml':
        use_ai = False
    else:  # ai 或 hybrid
        # 如果检测到敏感信息，强制使用ML
        if privacy_info.get("sensitive_detected", False):
            use_ai = False
            privacy_info["local_analysis_used"] = True
        else:
            use_ai = True if AI_AVAILABLE else False
    
    # 如果使用AI，对文本进行脱敏处理（保留上下文用于分析）
    analysis_text = original_text
    if use_ai and PRIVACY_AVAILABLE and privacy_info.get("sensitive_detected", False):
        # 发送到AI前进行脱敏（保留上下文）
        analysis_text, desensitization_log = privacy_utils.desensitize_text(
            original_text, 
            keep_context=True  # 保留上下文用于分析
        )
        privacy_info["desensitization_applied"] = len(desensitization_log.get('replacements', [])) > 0
    else:
        # 使用原始文本进行分析
        analysis_text = original_text
    
    # 情感分析
    sentiment_result = analyze_sentiment(analysis_text, use_ai=use_ai)
    
    # 关键词提取（使用分析文本，可能是脱敏后的）
    keywords = extract_keywords(analysis_text, topK=10, use_ai=use_ai if use_ai is not None else (AI_AVAILABLE and ai_analyzer.ai_enabled if AI_AVAILABLE else False))
    
    # 如果AI分析返回了关键词，使用AI的关键词
    # 判断是否使用了AI（通过查看sentiment_result的method字段）
    if sentiment_result.get('keywords') and 'ai' in sentiment_result.get('method', ''):
        keywords = [{"word": kw, "weight": 0.9 - (i * 0.1)} 
                   for i, kw in enumerate(sentiment_result['keywords'][:10])]
    
    # 返回响应（不包含原始敏感信息）
    # 如果进行了脱敏，返回脱敏后的文本
    response_text = analysis_text if privacy_info.get("desensitization_applied", False) else original_text
    
    response = {
        "text": response_text,  # 如果脱敏了，返回脱敏后的文本
        "sentiment": sentiment_result,
        "keywords": keywords,
        "analysis_info": {
            "method": sentiment_result.get("method", "ml"),
            "ai_enabled": AI_AVAILABLE and ai_analyzer.ai_enabled if AI_AVAILABLE else False,
            "mode": analysis_mode,
            "confidence": sentiment_result.get("confidence", 0.5),
            "human_review_needed": sentiment_result.get("human_review_needed", False),
            "review_status": sentiment_result.get("review_status", "needs_review"),
            "review_reason": sentiment_result.get("review_reason", ""),
            "confidence_thresholds": sentiment_result.get("confidence_thresholds", {
                "min": HUMAN_REVIEW_MIN_THRESHOLD,
                "max": HUMAN_REVIEW_MAX_THRESHOLD
            })
        },
        "privacy_info": privacy_info  # 隐私保护信息
    }
    
    # 如果AI分析提供了原因，添加到响应中
    if sentiment_result.get('reason'):
        response['analysis_reason'] = sentiment_result['reason']
    
    # 如果AI分析提供了负面部分和建议，添加到响应中
    if sentiment_result.get('negative_parts'):
        response['negative_parts'] = sentiment_result['negative_parts']
    if sentiment_result.get('suggestions'):
        response['suggestions'] = sentiment_result['suggestions']
    
    return jsonify(response)

@app.route('/api/analyze-batch', methods=['POST'])
def analyze_batch():
    """批量分析评论"""
    data = request.json
    comments = data.get('comments', [])
    analysis_mode = data.get('analysis_mode', 'auto')
    industry = data.get('industry', None)  # 行业类型  # auto, ai, ml, hybrid
    
    if not comments:
        return jsonify({"error": "评论列表不能为空"}), 400
    
    # 根据模式选择分析方法
    # auto模式：传递None让analyze_sentiment自动决定
    # ai/hybrid模式：强制使用AI
    # ml模式：强制使用ML
    if analysis_mode == 'auto':
        base_use_ai = None  # 让analyze_sentiment自动决定
    elif analysis_mode == 'ml':
        base_use_ai = False
    else:  # ai 或 hybrid
        base_use_ai = True if AI_AVAILABLE else False
    
    # 隐私保护：批量检测敏感信息
    batch_privacy_info = {
        "sensitive_detected_count": 0,
        "local_analysis_count": 0,
        "desensitization_applied_count": 0
    }
    
    # 批量分析时，如果选择AI模式且评论数量多，可以智能切换
    use_ai = base_use_ai
    # 如果评论太多，为了性能考虑，可以部分使用AI
    # 只有当use_ai为True时才考虑部分使用
    if use_ai is True and len(comments) > 10:
        # 前5条使用AI，其余使用ML
        use_ai_partial = True
    else:
        use_ai_partial = False
    
    # 判断是否使用批量AI分析
    use_batch_ai = (
        use_ai is True and 
        AI_AVAILABLE and 
        ai_analyzer.ai_enabled and 
        ai_analyzer.batch_enabled and 
        len(comments) > 1 and
        not use_ai_partial  # 部分使用AI时不启用批量模式
    )
    
    results = []
    all_keywords = []
    ai_count = 0
    ml_count = 0
    
    # 如果启用批量AI分析，先批量处理需要AI的评论
    ai_results_map = {}  # 存储批量分析结果
    if use_batch_ai:
        # 收集需要AI分析的评论（包括脱敏处理）
        ai_comments = []
        ai_indices = []
        ai_analysis_texts = []  # 脱敏后的文本
        ml_indices = []
        
        for i, comment in enumerate(comments):
            original_text = comment.get('text', '')
            if not original_text:
                continue
            
            # 检查隐私保护
            should_use_local = False
            sensitive_detected = False
            if PRIVACY_AVAILABLE:
                sensitive_info = privacy_utils.detect_sensitive_info(original_text)
                sensitive_detected = any(sensitive_info.values())
                should_use_local = privacy_utils.should_use_local_analysis(original_text, industry)
            
            if should_use_local:
                ml_indices.append(i)
            else:
                # 需要AI分析，进行脱敏处理
                analysis_text = original_text
                if sensitive_detected:
                    analysis_text, _ = privacy_utils.desensitize_text(original_text, keep_context=True)
                
                ai_comments.append({
                    'index': i,
                    'original': original_text,
                    'analysis': analysis_text,
                    'sensitive_detected': sensitive_detected
                })
                ai_indices.append(i)
                ai_analysis_texts.append(analysis_text)
        
        # 批量AI分析
        if ai_analysis_texts:
            print(f"[批量分析] 使用批量AI分析 {len(ai_analysis_texts)} 条评论")
            batch_results = ai_analyzer.analyze_sentiment_batch_with_ai(
                ai_analysis_texts, 
                batch_size=ai_analyzer.batch_size
            )
            
            # 将批量结果映射回原始位置，并保存评论信息
            for comment_info, result in zip(ai_comments, batch_results):
                if result:
                    # 保存评论的原始信息和脱敏信息
                    result['_comment_info'] = comment_info
                    ai_results_map[comment_info['index']] = result
        
        # 处理ML分析的评论
        if ml_indices:
            print(f"[批量分析] 使用ML分析 {len(ml_indices)} 条评论")
            for idx in ml_indices:
                comment = comments[idx]
                original_text = comment.get('text', '')
                ml_result = analyze_sentiment_ml(original_text)
                # 添加人机协同信息
                confidence = ml_result.get('confidence', 0.5)
                review_status = check_human_review_needed(confidence)
                ml_result['human_review_needed'] = review_status['needs_review']
                ml_result['review_status'] = review_status['status']
                ml_result['review_reason'] = review_status['reason']
                ml_result['confidence_thresholds'] = {
                    "min": HUMAN_REVIEW_MIN_THRESHOLD,
                    "max": HUMAN_REVIEW_MAX_THRESHOLD
                }
                ai_results_map[idx] = ml_result
    
    # 统一处理所有评论（兼容原有逻辑和批量结果）
    for i, comment in enumerate(comments):
        original_text = comment.get('text', '')
        if not original_text:
            continue
        
        # 隐私保护：检测敏感信息
        comment_privacy_info = {
            "sensitive_detected": False,
            "local_analysis_used": False,
            "desensitization_applied": False
        }
        
        if PRIVACY_AVAILABLE:
            sensitive_info = privacy_utils.detect_sensitive_info(original_text)
            comment_privacy_info["sensitive_detected"] = any(sensitive_info.values())
            
            if comment_privacy_info["sensitive_detected"]:
                batch_privacy_info["sensitive_detected_count"] += 1
            
            # 判断是否应该使用本地分析
            should_use_local = privacy_utils.should_use_local_analysis(original_text, industry)
            
            # 如果包含敏感信息或特殊行业，强制使用本地分析
            if should_use_local:
                current_use_ai = False
                comment_privacy_info["local_analysis_used"] = True
                batch_privacy_info["local_analysis_count"] += 1
            else:
                # 决定是否使用AI
                if use_ai_partial and i >= 5:
                    current_use_ai = False
                else:
                    current_use_ai = use_ai
        else:
            # 决定是否使用AI
            if use_ai_partial and i >= 5:
                current_use_ai = False
            else:
                current_use_ai = use_ai
        
        # 如果使用AI，对文本进行脱敏处理
        analysis_text = original_text
        if current_use_ai and PRIVACY_AVAILABLE and comment_privacy_info.get("sensitive_detected", False):
            # 发送到AI前进行脱敏（保留上下文）
            analysis_text, desensitization_log = privacy_utils.desensitize_text(
                original_text, 
                keep_context=True
            )
            comment_privacy_info["desensitization_applied"] = len(desensitization_log.get('replacements', [])) > 0
            if comment_privacy_info["desensitization_applied"]:
                batch_privacy_info["desensitization_applied_count"] += 1
        
        # 情感分析
        # 如果使用了批量AI分析，直接使用批量结果
        if use_batch_ai and i in ai_results_map:
            sentiment_result = ai_results_map[i].copy()  # 复制结果，避免修改原始数据
            
            # 获取评论信息（用于脱敏处理）
            comment_info = sentiment_result.pop('_comment_info', None)
            if comment_info:
                # 如果脱敏了，使用脱敏后的文本
                analysis_text = comment_info['analysis']
                comment_privacy_info["sensitive_detected"] = comment_info.get('sensitive_detected', False)
                if comment_info.get('sensitive_detected', False):
                    comment_privacy_info["desensitization_applied"] = True
                    batch_privacy_info["desensitization_applied_count"] += 1
                    if not comment_privacy_info.get("sensitive_detected"):
                        batch_privacy_info["sensitive_detected_count"] += 1
            
            # 添加人机协同信息（如果还没有）
            if 'human_review_needed' not in sentiment_result:
                confidence = sentiment_result.get('confidence', 0.5)
                review_status = check_human_review_needed(confidence)
                sentiment_result['human_review_needed'] = review_status['needs_review']
                sentiment_result['review_status'] = review_status['status']
                sentiment_result['review_reason'] = review_status['reason']
                sentiment_result['confidence_thresholds'] = {
                    "min": HUMAN_REVIEW_MIN_THRESHOLD,
                    "max": HUMAN_REVIEW_MAX_THRESHOLD
                }
        else:
            # 使用原有的单条分析逻辑
            sentiment_result = analyze_sentiment(analysis_text, use_ai=current_use_ai)
        
        # 统计分析方法
        if 'ai' in sentiment_result.get('method', ''):
            ai_count += 1
        else:
            ml_count += 1
        
        # 关键词提取（使用分析文本，可能是脱敏后的）
        keywords = extract_keywords(analysis_text, topK=5, use_ai=current_use_ai)
        
        # 如果AI分析返回了关键词，使用AI的关键词
        # 判断是否使用了AI（通过查看sentiment_result的method字段）
        if sentiment_result.get('keywords') and 'ai' in sentiment_result.get('method', ''):
            keywords = [{"word": kw, "weight": 0.9 - (j * 0.1)} 
                       for j, kw in enumerate(sentiment_result['keywords'][:5])]
        
        all_keywords.extend([kw['word'] for kw in keywords])
        
        # 返回文本（如果脱敏了，返回脱敏后的）
        response_text = analysis_text if comment_privacy_info.get("desensitization_applied", False) else original_text
        
        # 人机协同校验：获取复核状态
        review_status = sentiment_result.get("review_status", "needs_review")
        review_reason = sentiment_result.get("review_reason", "")
        
        results.append({
            "text": response_text,  # 如果脱敏了，返回脱敏后的文本
            "sentiment": sentiment_result,
            "keywords": keywords,
            "privacy_info": comment_privacy_info,  # 每条评论的隐私信息
            "human_review_needed": sentiment_result.get("human_review_needed", False),  # 是否需要人工复核（兼容旧版）
            "review_status": review_status,  # 复核状态：invalid | needs_review | auto_accepted
            "review_reason": review_reason,  # 状态说明
            "confidence": sentiment_result.get("confidence", 0.5)  # 置信度
        })
    
    # 统计所有关键词
    keyword_counter = Counter(all_keywords)
    top_keywords = [{"word": word, "count": count} 
                   for word, count in keyword_counter.most_common(20)]
    
    # 计算满意度评分
    satisfaction = calculate_satisfaction_score(comments)
    
    # 情感分布统计
    sentiment_dist = Counter([r['sentiment']['sentiment'] for r in results])
    
    # 人机协同统计：统计不同状态的评论数量
    human_review_count = sum(1 for r in results if r.get('review_status') == 'needs_review')
    invalid_count = sum(1 for r in results if r.get('review_status') == 'invalid')
    auto_accepted_count = sum(1 for r in results if r.get('review_status') == 'auto_accepted')
    human_review_rate = (human_review_count / len(results) * 100) if results else 0
    
    return jsonify({
        "results": results,
        "privacy_info": batch_privacy_info,  # 批量分析的隐私统计
        "statistics": {
            "satisfaction": satisfaction,
            "sentiment_distribution": dict(sentiment_dist),
            "top_keywords": top_keywords,
            "human_review": {
                "needed_count": human_review_count,
                "invalid_count": invalid_count,
                "auto_accepted_count": auto_accepted_count,
                "total_count": len(results),
                "review_rate": round(human_review_rate, 2)
            }
        },
        "analysis_info": {
            "mode": analysis_mode,
            "ai_enabled": AI_AVAILABLE and ai_analyzer.ai_enabled if AI_AVAILABLE else False,
            "ai_analysis_count": ai_count,
            "ml_analysis_count": ml_count,
            "total_comments": len(results),
            "confidence_thresholds": {
                "min": HUMAN_REVIEW_MIN_THRESHOLD,
                "max": HUMAN_REVIEW_MAX_THRESHOLD
            }
        }
    })

@app.route('/api/export-report', methods=['POST'])
def export_report():
    """
    导出分析报告（Excel格式）
    
    请求体应包含完整的批量分析结果数据
    """
    try:
        data = request.json
        if not data:
            return jsonify({"error": "请求数据不能为空"}), 400
        
        results = data.get('results', [])
        statistics = data.get('statistics', {})
        analysis_info = data.get('analysis_info', {})
        privacy_info = data.get('privacy_info', {})
        
        if not results:
            return jsonify({"error": "分析结果不能为空"}), 400
        
        # 创建Excel工作簿
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        wb = Workbook()
        
        # 删除默认工作表
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # ========== 工作表1：评论详情 ==========
        ws1 = wb.create_sheet("评论详情", 0)
        
        # 设置表头
        headers1 = [
            "序号", "评论内容", "情感倾向", "置信度", "复核状态", "复核原因",
            "关键词", "负面部分", "改进建议", "分析原因", "分析方法"
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_idx, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 情感标签映射
        sentiment_map = {
            "strongly_negative": "强烈负面",
            "weakly_negative": "轻微负面",
            "neutral": "中性",
            "weakly_positive": "轻微正面",
            "strongly_positive": "强烈正面",
            "positive": "正面",
            "negative": "负面"
        }
        
        # 复核状态映射
        review_status_map = {
            "invalid": "无效/待观察",
            "needs_review": "需要复核",
            "auto_accepted": "自动采纳"
        }
        
        # 写入数据
        for row_idx, result in enumerate(results, 2):
            sentiment_data = result.get('sentiment', {})
            sentiment_label = sentiment_data.get('sentiment', 'neutral')
            sentiment_cn = sentiment_map.get(sentiment_label, sentiment_label)
            
            keywords = result.get('keywords', [])
            keywords_str = ', '.join([kw.get('word', kw) if isinstance(kw, dict) else kw for kw in keywords[:5]])
            
            negative_parts = sentiment_data.get('negative_parts', [])
            negative_parts_str = '; '.join(negative_parts) if negative_parts else '无'
            
            suggestions = sentiment_data.get('suggestions', [])
            suggestions_str = '; '.join(suggestions) if suggestions else '无'
            
            review_status = result.get('review_status', 'needs_review')
            review_status_cn = review_status_map.get(review_status, review_status)
            
            # 安全获取置信度（确保是数字类型）
            confidence_value = result.get('confidence', 0)
            if isinstance(confidence_value, dict):
                confidence_value = confidence_value.get('value', 0) if 'value' in confidence_value else 0
            try:
                confidence_str = f"{float(confidence_value):.2%}"
            except (ValueError, TypeError):
                confidence_str = "0.00%"
            
            # 获取分析方法（优先从sentiment_data获取，如果没有则尝试从result直接获取）
            method = sentiment_data.get('method', '')
            if not method:
                # 如果sentiment_data中没有method，尝试从result中获取（兼容某些情况）
                method = result.get('method', 'ml')
            # 如果还是没有，根据是否有AI相关字段判断
            if not method or method == 'ml':
                # 检查是否有AI分析的特征字段
                if sentiment_data.get('keywords') or sentiment_data.get('negative_parts') or sentiment_data.get('suggestions'):
                    # 如果有这些字段，可能是AI分析但没有method字段，尝试推断
                    if 'ai' in str(sentiment_data.get('reason', '')).lower() or 'ai' in str(result.get('analysis_reason', '')).lower():
                        method = 'ai_unknown'
                    else:
                        method = 'ml'
                else:
                    method = 'ml'
            
            # 格式化方法显示名称
            method_display_map = {
                'ai_deepseek': 'AI (DeepSeek)',
                'ai_openai': 'AI (OpenAI)',
                'ai_deepseek_batch': 'AI (DeepSeek批量)',
                'ai_openai_batch': 'AI (OpenAI批量)',
                'hybrid_ai': '混合 (AI优先)',
                'hybrid_ml': '混合 (ML优先)',
                'ml': 'ML (机器学习)',
                'ai_unknown': 'AI (未知来源)'
            }
            method_display = method_display_map.get(method, method)
            
            row_data = [
                row_idx - 1,  # 序号
                result.get('text', ''),  # 评论内容
                sentiment_cn,  # 情感倾向
                confidence_str,  # 置信度
                review_status_cn,  # 复核状态
                result.get('review_reason', ''),  # 复核原因
                keywords_str,  # 关键词
                negative_parts_str,  # 负面部分
                suggestions_str,  # 改进建议
                sentiment_data.get('reason', ''),  # 分析原因
                method_display  # 分析方法
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 调整列宽
        column_widths = [8, 40, 12, 10, 12, 15, 25, 30, 30, 30, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 冻结首行
        ws1.freeze_panes = 'A2'
        
        # ========== 工作表2：统计分析 ==========
        ws2 = wb.create_sheet("统计分析", 1)
        
        # 满意度评分
        satisfaction = statistics.get('satisfaction', 0)
        # 处理satisfaction可能是字典的情况
        if isinstance(satisfaction, dict):
            satisfaction = satisfaction.get('score', 0)
        try:
            satisfaction_value = float(satisfaction)
        except (ValueError, TypeError):
            satisfaction_value = 0.0
        ws2['A1'] = "满意度评分"
        ws2['B1'] = f"{satisfaction_value:.2f}/5.0"
        ws2['A1'].font = Font(bold=True, size=12)
        ws2['B1'].font = Font(size=12)
        
        # 情感分布
        ws2['A3'] = "情感分布"
        ws2['A3'].font = Font(bold=True, size=12)
        
        sentiment_dist = statistics.get('sentiment_distribution', {})
        ws2['A4'] = "情感类型"
        ws2['B4'] = "数量"
        ws2['C4'] = "占比"
        
        # 设置表头样式
        for col in ['A4', 'B4', 'C4']:
            cell = ws2[col]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        total_count = sum(sentiment_dist.values())
        row = 5
        for sentiment, count in sentiment_dist.items():
            sentiment_cn = sentiment_map.get(sentiment, sentiment)
            percentage = (count / total_count * 100) if total_count > 0 else 0
            ws2[f'A{row}'] = sentiment_cn
            ws2[f'B{row}'] = count
            ws2[f'C{row}'] = f"{percentage:.2f}%"
            for col in ['A', 'B', 'C']:
                ws2[f'{col}{row}'].border = border
            row += 1
        
        # 关键词统计
        ws2[f'A{row+1}'] = "关键词统计（Top 20）"
        ws2[f'A{row+1}'].font = Font(bold=True, size=12)
        
        top_keywords = statistics.get('top_keywords', [])
        ws2[f'A{row+2}'] = "关键词"
        ws2[f'B{row+2}'] = "出现次数"
        
        # 设置表头样式
        for col in [f'A{row+2}', f'B{row+2}']:
            cell = ws2[col]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        keyword_row = row + 3
        for keyword_data in top_keywords[:20]:
            word = keyword_data.get('word', keyword_data) if isinstance(keyword_data, dict) else keyword_data
            count = keyword_data.get('count', 0) if isinstance(keyword_data, dict) else 0
            ws2[f'A{keyword_row}'] = word
            ws2[f'B{keyword_row}'] = count
            for col in ['A', 'B']:
                ws2[f'{col}{keyword_row}'].border = border
            keyword_row += 1
        
        # 调整列宽
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 12
        
        # ========== 工作表3：人工复核统计 ==========
        ws3 = wb.create_sheet("人工复核统计", 2)
        
        human_review = statistics.get('human_review', {})
        ws3['A1'] = "人工复核统计"
        ws3['A1'].font = Font(bold=True, size=14)
        
        ws3['A3'] = "项目"
        ws3['B3'] = "数量"
        ws3['C3'] = "占比"
        
        # 设置表头样式
        for col in ['A3', 'B3', 'C3']:
            cell = ws3[col]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        review_data = [
            ("需要复核", human_review.get('needed_count', 0)),
            ("无效/待观察", human_review.get('invalid_count', 0)),
            ("自动采纳", human_review.get('auto_accepted_count', 0)),
            ("总计", human_review.get('total_count', 0))
        ]
        
        total_review = human_review.get('total_count', 0)
        row = 4
        for label, count in review_data:
            percentage = (count / total_review * 100) if total_review > 0 else 0
            ws3[f'A{row}'] = label
            ws3[f'B{row}'] = count
            ws3[f'C{row}'] = f"{percentage:.2f}%"
            for col in ['A', 'B', 'C']:
                ws3[f'{col}{row}'].border = border
            row += 1
        
        ws3['A8'] = "复核率"
        review_rate = human_review.get('review_rate', 0)
        try:
            review_rate_value = float(review_rate)
        except (ValueError, TypeError):
            review_rate_value = 0.0
        ws3['B8'] = f"{review_rate_value:.2f}%"
        ws3['A8'].font = Font(bold=True)
        ws3['B8'].font = Font(bold=True)
        
        # 调整列宽
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 15
        ws3.column_dimensions['C'].width = 12
        
        # ========== 工作表4：分析信息 ==========
        ws4 = wb.create_sheet("分析信息", 3)
        
        ws4['A1'] = "分析配置信息"
        ws4['A1'].font = Font(bold=True, size=14)
        
        # 安全获取置信度阈值
        confidence_thresholds = analysis_info.get('confidence_thresholds', {})
        min_threshold = confidence_thresholds.get('min', 0.5) if isinstance(confidence_thresholds, dict) else 0.5
        max_threshold = confidence_thresholds.get('max', 0.85) if isinstance(confidence_thresholds, dict) else 0.85
        try:
            min_threshold_value = float(min_threshold)
            max_threshold_value = float(max_threshold)
        except (ValueError, TypeError):
            min_threshold_value = 0.5
            max_threshold_value = 0.85
        
        analysis_config = [
            ("分析模式", analysis_info.get('mode', 'auto')),
            ("AI分析启用", "是" if analysis_info.get('ai_enabled', False) else "否"),
            ("AI分析数量", analysis_info.get('ai_analysis_count', 0)),
            ("ML分析数量", analysis_info.get('ml_analysis_count', 0)),
            ("总评论数", analysis_info.get('total_comments', 0)),
            ("置信度阈值（最小）", f"{min_threshold_value:.2f}"),
            ("置信度阈值（最大）", f"{max_threshold_value:.2f}"),
        ]
        
        row = 3
        for label, value in analysis_config:
            ws4[f'A{row}'] = label
            ws4[f'B{row}'] = value
            ws4[f'A{row}'].font = Font(bold=True)
            ws4[f'A{row}'].border = border
            ws4[f'B{row}'].border = border
            row += 1
        
        # 隐私保护信息
        if privacy_info:
            row += 1
            ws4[f'A{row}'] = "隐私保护统计"
            ws4[f'A{row}'].font = Font(bold=True, size=12)
            
            privacy_config = [
                ("检测到敏感信息", privacy_info.get('sensitive_detected_count', 0)),
                ("使用本地分析", privacy_info.get('local_analysis_count', 0)),
                ("应用脱敏处理", privacy_info.get('desensitization_applied_count', 0)),
            ]
            
            row += 2
            for label, value in privacy_config:
                ws4[f'A{row}'] = label
                ws4[f'B{row}'] = value
                ws4[f'A{row}'].border = border
                ws4[f'B{row}'].border = border
                row += 1
        
        # 生成时间
        row += 1
        ws4[f'A{row}'] = "报告生成时间"
        ws4[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws4[f'A{row}'].font = Font(bold=True)
        
        # 调整列宽
        ws4.column_dimensions['A'].width = 25
        ws4.column_dimensions['B'].width = 30
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"评论分析报告_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"[导出报告] 错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"导出报告失败: {str(e)}"}), 500

@app.route('/api/health', methods=['GET'])
def health():
    """健康检查"""
    return jsonify({
        "status": "ok",
        "message": "服务运行正常",
        "ai_available": AI_AVAILABLE,
        "ai_enabled": AI_AVAILABLE and ai_analyzer.ai_enabled if AI_AVAILABLE else False
    })

@app.route('/api/config', methods=['GET'])
def get_config():
    """获取分析配置信息"""
    return jsonify({
        "analysis_modes": {
            "auto": "自动选择（优先AI，失败时使用ML）",
            "ai": "仅使用AI分析",
            "ml": "仅使用机器学习模型",
            "hybrid": "混合模式（AI+ML结合）"
        },
        "current_status": {
            "ai_available": AI_AVAILABLE,
            "ai_enabled": AI_AVAILABLE and ai_analyzer.ai_enabled if AI_AVAILABLE else False,
            "ml_available": ML_AVAILABLE and ml_analyzer is not None if ML_AVAILABLE else False,
            "ml_trained": ml_analyzer.is_trained if (ML_AVAILABLE and ml_analyzer) else False,
            "ml_training_data_count": ml_analyzer.get_model_info()['training_data_count'] if (ML_AVAILABLE and ml_analyzer) else 0
        },
        "human_review_config": {
            "enabled": True,
            "confidence_thresholds": {
                "min": HUMAN_REVIEW_MIN_THRESHOLD,
                "max": HUMAN_REVIEW_MAX_THRESHOLD
            },
            "description": f"置信度规则：< {HUMAN_REVIEW_MIN_THRESHOLD} 标记为无效/待观察，{HUMAN_REVIEW_MIN_THRESHOLD}-{HUMAN_REVIEW_MAX_THRESHOLD} 需要人工复核，> {HUMAN_REVIEW_MAX_THRESHOLD} 自动采纳",
            "auto_add_to_training": AUTO_ADD_TO_TRAINING,
            "auto_retrain_on_update": AUTO_RETRAIN_ON_UPDATE
        }
    })

@app.route('/api/retrain', methods=['POST'])
def retrain_model():
    """
    重新训练模型（使用最新的训练数据）
    """
    global ml_analyzer, training_data
    
    if not ML_AVAILABLE:
        return jsonify({"error": "ML分析模块未启用"}), 400
    
    try:
        # 重新加载训练数据
        training_data = load_training_data()
        
        # 重新训练模型
        ml_analyzer = create_ml_analyzer(training_data)
        
        # 获取模型信息
        model_info = ml_analyzer.get_model_info()
        
        return jsonify({
            "success": True,
            "message": "模型重新训练成功",
            "training_data_count": len(training_data),
            "model_info": model_info
        })
    except Exception as e:
        return jsonify({"error": f"模型训练失败: {str(e)}"}), 500

@app.route('/api/human-review', methods=['POST'])
def submit_human_review():
    """
    提交人工复核结果
    用于人机协同：人工复核低置信度的评论，并反馈结果用于模型优化
    
    Request Body:
    {
        "comment_id": "评论ID（可选，用于追踪）",
        "text": "评论文本",
        "reviewed_sentiment": "strongly_negative/weakly_negative/neutral/weakly_positive/strongly_positive (也兼容旧标签positive/negative/neutral)",  # 人工复核的情感
        "reviewed_confidence": 0.0-1.0,  # 人工复核的置信度（可选）
        "review_notes": "复核备注（可选）"
    }
    """
    data = request.json
    text = data.get('text', '')
    reviewed_sentiment = data.get('reviewed_sentiment', '')
    reviewed_confidence = data.get('reviewed_confidence', 1.0)
    comment_id = data.get('comment_id', None)
    review_notes = data.get('review_notes', '')
    
    if not text:
        return jsonify({"error": "评论文本不能为空"}), 400
    
    if reviewed_sentiment not in ['positive', 'negative', 'neutral']:
        return jsonify({"error": "复核情感必须是positive、negative或neutral之一"}), 400
    
    # 生成评论的唯一标识（如果没有提供comment_id）
    if not comment_id:
        import hashlib
        comment_id = hashlib.md5(text.encode('utf-8')).hexdigest()
    
    # 获取模型初筛结果（用于对比）
    model_result = analyze_sentiment(text)
    model_sentiment = model_result.get('sentiment', '')
    model_confidence = model_result.get('confidence', 0.5)
    
    # 判断是否需要人工复核（如果模型置信度低，人工复核更有价值）
    review_status = check_human_review_needed(model_confidence)
    was_low_confidence = review_status['needs_review'] or review_status['status'] == 'invalid'
    
    # 判断人工复核结果与模型结果是否一致
    is_consistent = (model_sentiment == reviewed_sentiment)
    
    # 存储人工复核结果（实际应用中应使用数据库）
    human_review_results[comment_id] = {
        "text": text,
        "model_sentiment": model_sentiment,
        "model_confidence": model_confidence,
        "reviewed_sentiment": reviewed_sentiment,
        "reviewed_confidence": reviewed_confidence,
        "is_consistent": is_consistent,
        "was_low_confidence": was_low_confidence,
        "review_notes": review_notes,
        "review_timestamp": __import__('datetime').datetime.now().isoformat()
    }
    
    # 自动添加到训练数据（如果启用）
    added_to_training = False
    retrained = False
    if AUTO_ADD_TO_TRAINING:
        added_to_training = add_to_training_data(text, reviewed_sentiment)
        
        # 如果成功添加到训练数据，且启用了自动重训练，则重新训练模型
        if added_to_training and AUTO_RETRAIN_ON_UPDATE and ML_AVAILABLE:
            try:
                global ml_analyzer, training_data
                # 重新加载训练数据
                training_data = load_training_data()
                # 重新训练模型
                ml_analyzer = create_ml_analyzer(training_data)
                retrained = True
                print(f"[自动重训练] 模型已自动重新训练，使用 {len(training_data)} 条数据")
            except Exception as e:
                print(f"[自动重训练] 自动重训练失败: {e}")
    
    response_data = {
        "success": True,
        "message": "人工复核结果已提交",
        "comment_id": comment_id,
        "comparison": {
            "model_sentiment": model_sentiment,
            "model_confidence": model_confidence,
            "reviewed_sentiment": reviewed_sentiment,
            "reviewed_confidence": reviewed_confidence,
            "is_consistent": is_consistent,
            "was_low_confidence": was_low_confidence
        },
        "training": {
            "added_to_training": added_to_training,
            "auto_retrained": retrained
        }
    }
    
    if added_to_training:
        response_data["note"] = "人工复核结果已添加到训练数据" + ("，模型已自动重新训练" if retrained else "，可手动调用 /api/retrain 重新训练模型")
    else:
        response_data["note"] = "人工复核结果已记录，可用于后续模型优化和训练数据增强"
    
    return jsonify(response_data)

@app.route('/api/human-review/stats', methods=['GET'])
def get_human_review_stats():
    """
    获取人工复核统计信息
    """
    if not human_review_results:
        return jsonify({
            "total_reviews": 0,
            "message": "暂无人工复核数据"
        })
    
    total_reviews = len(human_review_results)
    consistent_count = sum(1 for r in human_review_results.values() if r.get('is_consistent', False))
    low_confidence_count = sum(1 for r in human_review_results.values() if r.get('was_low_confidence', False))
    
    # 统计情感分布
    sentiment_dist = Counter([r.get('reviewed_sentiment', '') for r in human_review_results.values()])
    
    return jsonify({
        "total_reviews": total_reviews,
        "statistics": {
            "consistent_count": consistent_count,
            "inconsistent_count": total_reviews - consistent_count,
            "consistency_rate": round(consistent_count / total_reviews * 100, 2) if total_reviews > 0 else 0,
            "low_confidence_count": low_confidence_count,
            "low_confidence_rate": round(low_confidence_count / total_reviews * 100, 2) if total_reviews > 0 else 0,
            "reviewed_sentiment_distribution": dict(sentiment_dist)
        },
        "confidence_thresholds": {
            "min": HUMAN_REVIEW_MIN_THRESHOLD,
            "max": HUMAN_REVIEW_MAX_THRESHOLD
        }
    })

@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    """
    上传Excel文件并解析评论数据
    
    支持格式：
    - .xlsx (Excel 2007+)
    - .xls (Excel 97-2003)
    
    Excel文件应包含以下列（至少需要"评论"列）：
    - 评论（必需）：评论文本内容
    - 其他列会被忽略
    """
    if 'file' not in request.files:
        return jsonify({"error": "未找到上传的文件"}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"error": "未选择文件"}), 400
    
    # 检查文件扩展名（使用原始文件名，不经过secure_filename处理）
    original_filename = file.filename or ''
    filename_lower = original_filename.lower().strip()
    
    # 调试信息
    print(f"[Excel导入] 原始文件名: {original_filename}")
    print(f"[Excel导入] 小写文件名: {filename_lower}")
    print(f"[Excel导入] 文件MIME类型: {file.content_type}")
    
    # 检查文件扩展名
    has_valid_extension = filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls')
    
    # 也检查MIME类型（作为备用验证）
    valid_mime_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',  # .xls
        'application/excel',
        'application/x-excel',
        'application/x-msexcel'
    ]
    has_valid_mime = file.content_type in valid_mime_types if file.content_type else False
    
    if not has_valid_extension and not has_valid_mime:
        detected_ext = filename_lower.split('.')[-1] if '.' in filename_lower else '无'
        return jsonify({
            "error": f"不支持的文件格式，请上传 .xlsx 或 .xls 文件。\n当前文件: {original_filename}\n检测到的扩展名: {detected_ext}\nMIME类型: {file.content_type or '未知'}"
        }), 400
    
    # 使用secure_filename处理文件名（仅用于安全，不影响扩展名检查）
    filename = secure_filename(original_filename) if original_filename else 'upload.xlsx'
    
    try:
        # 读取Excel文件（根据原始文件名或MIME类型判断引擎）
        if filename_lower.endswith('.xlsx'):
            use_openpyxl = True
        elif filename_lower.endswith('.xls'):
            use_openpyxl = False
        elif file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            use_openpyxl = True
        else:
            # 默认尝试使用openpyxl（适用于.xlsx）
            use_openpyxl = True
        
        print(f"[Excel导入] 使用引擎: {'openpyxl' if use_openpyxl else 'xlrd'}")
        df = pd.read_excel(file, engine='openpyxl' if use_openpyxl else None)
        
        if df.empty:
            return jsonify({"error": "Excel文件为空"}), 400
        
        # 查找评论列（支持多种可能的列名）
        comment_column = None
        possible_names = ['评论', 'comment', 'comments', '内容', '文本', 'text', 'content', '评论文本']
        
        for col in df.columns:
            if str(col).strip() in possible_names:
                comment_column = col
                break
        
        # 如果没找到，尝试使用第一列
        if comment_column is None:
            comment_column = df.columns[0]
            print(f"[Excel导入] 未找到标准评论列，使用第一列: {comment_column}")
        
        # 提取评论数据
        comments = []
        for idx, row in df.iterrows():
            comment_text = str(row[comment_column]).strip()
            if comment_text and comment_text.lower() != 'nan' and comment_text != '':
                comments.append({
                    "text": comment_text
                })
        
        if not comments:
            return jsonify({"error": "Excel文件中未找到有效的评论数据"}), 400
        
        return jsonify({
            "success": True,
            "message": f"成功导入 {len(comments)} 条评论",
            "comments": comments,
            "total": len(comments)
        })
        
    except Exception as e:
        print(f"[Excel导入] 错误: {str(e)}")
        return jsonify({"error": f"解析Excel文件失败: {str(e)}"}), 500

@app.route('/api/download-template', methods=['GET'])
def download_template():
    """
    下载Excel模板文件
    模板包含示例评论数据，用户可以参考格式导入自己的数据
    """
    try:
        # 创建示例数据
        sample_data = {
            '评论': [
                '这个商品质量很好，非常满意！',
                '物流很快，包装精美，五星好评！',
                '质量不错，价格实惠',
                '商品质量差，不推荐',
                '一般般，还可以',
                '非常喜欢，下次还会购买',
                '不满意，退货了',
                '性价比很高，推荐购买'
            ]
        }
        
        df = pd.DataFrame(sample_data)
        
        # 创建Excel文件在内存中
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='评论数据')
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='评论数据导入模板.xlsx'
        )
        
    except Exception as e:
        print(f"[模板下载] 错误: {str(e)}")
        return jsonify({"error": f"生成模板文件失败: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, port=8080)

